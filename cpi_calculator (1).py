"""
Australia Post — CPI Calculator
Data & Pricing Executive Tool  |  Version 3.0
Source: ABS 6401.0 Consumer Price Index, Australia (Monthly + Quarterly)
Base period: September 2025 = 100.00
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

from abs_cpi_parser import load_abs_file, get_city_df, calc_custom_change, CITIES

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CPI Calculator | Australia Post",
    page_icon="📮",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── STYLES ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=IBM+Plex+Sans:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }

.ap-header {
    background: #C8102E;
    padding: 24px 32px;
    border-radius: 10px;
    margin-bottom: 24px;
    display: flex; align-items: center; gap: 18px;
    border-left: 6px solid #8B0020;
}
.ap-header h1 { color: white; font-size: 1.75rem; font-weight: 700; margin: 0; letter-spacing: -0.03em; }
.ap-header .sub { color: rgba(255,255,255,0.75); font-size: 0.82rem; margin-top: 5px; font-weight: 300; }
.ap-logo { font-size: 2.5rem; }

.kpi-wrap { display: flex; gap: 12px; margin-bottom: 24px; flex-wrap: wrap; }
.kpi { background: white; border: 1px solid #e5e5e5; border-radius: 8px;
       padding: 18px 20px; flex: 1; min-width: 130px; box-shadow: 0 1px 4px rgba(0,0,0,0.05); }
.kpi-label { font-size: 0.68rem; font-weight: 700; text-transform: uppercase;
             letter-spacing: 0.1em; color: #999; margin-bottom: 8px; }
.kpi-val { font-family: 'IBM Plex Mono', monospace; font-size: 1.9rem; font-weight: 500; color: #111; line-height: 1; }
.kpi-sub { font-size: 0.72rem; color: #bbb; margin-top: 5px; }
.kpi-red .kpi-val { color: #C8102E; }
.kpi-green .kpi-val { color: #15803d; }
.kpi-blue .kpi-val { color: #1D4ED8; }

.signal-box { border-radius: 8px; padding: 14px 20px; margin: 16px 0; display: flex; gap: 14px; align-items: flex-start; }
.signal-icon { font-size: 1.6rem; line-height: 1; }
.signal-title { font-weight: 700; font-size: 0.95rem; color: #111; margin-bottom: 4px; }
.signal-desc { font-size: 0.85rem; color: #444; line-height: 1.5; }
.signal-formula { font-family: 'IBM Plex Mono', monospace; font-size: 0.75rem; color: #888; margin-top: 6px; }

.series-pill {
    display: inline-block; padding: 3px 10px; border-radius: 20px;
    font-size: 0.72rem; font-weight: 700; margin-right: 6px; vertical-align: middle;
}
.pill-monthly { background: #FEE2E2; color: #C8102E; }
.pill-quarterly { background: #DBEAFE; color: #1D4ED8; }

.sec-title { font-size: 0.68rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.12em;
             color: #C8102E; border-bottom: 2px solid #C8102E; display: inline-block;
             padding-bottom: 4px; margin-bottom: 14px; }

.badge { display: inline-block; background: #FEE2E2; color: #C8102E; font-size: 0.7rem;
         font-weight: 700; padding: 2px 8px; border-radius: 20px; margin-left: 8px; vertical-align: middle; }
.badge-green { background: #DCFCE7; color: #15803d; }
.badge-yellow { background: #FEF9C3; color: #854D0E; }

.stDownloadButton > button {
    background: #C8102E !important; color: white !important;
    border: none !important; font-weight: 600 !important;
    border-radius: 6px !important; padding: 8px 18px !important; font-size: 0.85rem !important;
}
.stDownloadButton > button:hover { background: #9B0B23 !important; }
section[data-testid="stSidebar"] { background: #FAFAFA; border-right: 1px solid #EEE; }

section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"],
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p,
section[data-testid="stSidebar"] .stRadio label,
section[data-testid="stSidebar"] .stCheckbox label,
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span {
    color: #111111 !important;
}

.abs-note { background: #F0F7FF; border-left: 4px solid #3B82F6; padding: 10px 14px;
            border-radius: 0 6px 6px 0; font-size: 0.8rem; color: #1E3A5F; margin: 10px 0; }

.dual-header { background: linear-gradient(135deg, #FEF2F2 0%, #EFF6FF 100%);
               border: 1px solid #E5E7EB; border-radius: 8px; padding: 12px 18px;
               margin-bottom: 18px; display: flex; gap: 24px; align-items: center; }
.dual-header-item { display: flex; flex-direction: column; gap: 2px; }
.dual-header-label { font-size: 0.68rem; font-weight: 700; text-transform: uppercase;
                     letter-spacing: 0.1em; color: #999; }
.dual-header-val { font-family: 'IBM Plex Mono', monospace; font-size: 1.1rem; font-weight: 600; }
.dual-monthly { color: #C8102E; }
.dual-quarterly { color: #1D4ED8; }
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTS ────────────────────────────────────────────────────────────────
MONTHLY_COLOR   = "#C8102E"
QUARTERLY_COLOR = "#1D4ED8"

# ─── DATA LOAD ───────────────────────────────────────────────────────────────

@st.cache_data
def load_bundled_data():
    try:
        with open("640101.xlsx", "rb") as f:
            return load_abs_file(f.read())
    except FileNotFoundError:
        return None

def try_load_upload(uploaded_file):
    try:
        return load_abs_file(uploaded_file.read())
    except Exception as e:
        st.error(f"Could not parse file: {e}\nPlease ensure this is an ABS 6401.0 Table 1 or Table 6 Time Series workbook.")
        return None

# ─── EXCEL EXPORT ────────────────────────────────────────────────────────────

def build_excel_report(city_df, city, start_period, end_period, calc, frequency="Monthly"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CPI Report"

    RED, BLUE       = "C8102E", "1D4ED8"
    LIGHT_RED       = "FCE8EC"
    LIGHT_BLUE      = "DBEAFE"
    GREY, WHITE     = "F5F5F5", "FFFFFF"
    DARK            = "1A1A1A"
    HDR_COLOR       = RED if frequency == "Monthly" else BLUE
    LIGHT_COLOR     = LIGHT_RED if frequency == "Monthly" else LIGHT_BLUE
    thin = Side(style="thin", color="E0E0E0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(cell, value, bg=None, fg="FFFFFF", sz=11, bold=True):
        cell.value = value
        cell.font = Font(name="Calibri", bold=bold, size=sz, color=fg)
        cell.fill = PatternFill("solid", fgColor=bg or HDR_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr

    ws.merge_cells("A1:G1")
    hdr_cell(ws["A1"], f"Australia Post — CPI Analysis Report ({frequency})", sz=14)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Region: {city}   |   Period: {start_period} → {end_period}   |   ABS 6401.0  (Base: Sep-2025 = 100)"
    ws["A2"].font = Font(name="Calibri", size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    kpis = [
        ("Start Index",  f"{calc['start_val']:.2f}",  "D"),
        ("End Index",    f"{calc['end_val']:.2f}",    "E"),
        ("Movement",     f"{calc['movement']:+.2f} pts", "F"),
        ("% Change",     f"{calc['pct_change']:+.2f}%",  "G"),
    ]
    for label, value, col in kpis:
        ws[f"{col}4"] = label
        ws[f"{col}4"].font = Font(name="Calibri", bold=True, size=8, color="888888")
        ws[f"{col}4"].alignment = Alignment(horizontal="center")
        ws[f"{col}5"] = value
        ws[f"{col}5"].font = Font(name="Calibri", bold=True, size=16,
                                  color=HDR_COLOR if "%" in value or "pts" in value else DARK)
        ws[f"{col}5"].fill = PatternFill("solid", fgColor=LIGHT_COLOR if "%" in value else GREY)
        ws[f"{col}5"].alignment = Alignment(horizontal="center")
        ws[f"{col}5"].border = bdr
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 28

    pct = calc['pct_change']
    signal = "🔴 Strong Review Recommended" if abs(pct) >= 4 else ("🟡 Review Warranted" if abs(pct) >= 2 else "🟢 Stable — Monitor Quarterly")
    ws.merge_cells("D6:G6")
    ws["D6"] = f"Pricing Signal: {signal}"
    ws["D6"].font = Font(name="Calibri", bold=True, size=10, color=HDR_COLOR)
    ws["D6"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[6].height = 22

    period_header = "Period"
    pop_label = "QoQ Change (%)" if frequency == "Quarterly" else "MoM Change (%)"
    headers = [period_header, "CPI Index", pop_label, "YoY Change (%)", "vs Start (%)", "Signal"]
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=8, column=col_i, value=h)
        hdr_cell(c, h, sz=9)
    ws.row_dimensions[8].height = 20

    mask = (city_df['Date'] >= pd.to_datetime(start_period.replace("Q1-","Jan-").replace("Q2-","Apr-").replace("Q3-","Jul-").replace("Q4-","Oct-"), format="%b-%Y" if "Q" not in start_period else None)) & \
           (city_df['Date'] <= pd.to_datetime(end_period.replace("Q1-","Jan-").replace("Q2-","Apr-").replace("Q3-","Jul-").replace("Q4-","Oct-"), format="%b-%Y" if "Q" not in end_period else None))
    # Simpler: filter by period string
    periods_in_range = city_df[(city_df['Period'] >= start_period) & (city_df['Period'] <= end_period)] if False else city_df[
        city_df['Period'].isin(city_df[
            (city_df['Date'] >= city_df[city_df['Period'] == start_period]['Date'].values[0]) &
            (city_df['Date'] <= city_df[city_df['Period'] == end_period]['Date'].values[0])
        ]['Period'])
    ]
    sub = periods_in_range.copy()
    sub['vs_start'] = ((sub['CPI_Index'] - calc['start_val']) / calc['start_val'] * 100).round(2)

    for row_i, (_, row) in enumerate(sub.iterrows(), 9):
        fill_c = "F9F9F9" if row_i % 2 == 0 else WHITE
        mom_val = row['MoM_Pct']
        yoy_val = row['YoY_Pct']
        vs_val  = row['vs_start']
        row_signal = "▲ Rising" if (mom_val or 0) > 0.5 else ("▼ Easing" if (mom_val or 0) < 0 else "→ Flat")
        vals = [row['Period'], row['CPI_Index'], mom_val, yoy_val, vs_val, row_signal]
        for col_i, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=col_i, value=v)
            c.font = Font(name="Calibri", size=9)
            c.alignment = Alignment(horizontal="center")
            c.fill = PatternFill("solid", fgColor=fill_c)
            c.border = bdr
            if col_i in (3, 4, 5) and isinstance(v, (int, float)) and v is not None:
                c.number_format = '+0.00;-0.00;0.00'
        ws.row_dimensions[row_i].height = 15

    for col, w in zip("ABCDEFG", [12, 11, 15, 14, 13, 12, 14]):
        ws.column_dimensions[get_column_letter(ord(col)-64)].width = w

    ws2 = wb.create_sheet("Raw Data")
    pop_col = "QoQ % Change" if frequency == "Quarterly" else "MoM % Change"
    raw_headers = ["Period", "CPI Index", pop_col, "YoY % Change"]
    for i, h in enumerate(raw_headers, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        c.fill = PatternFill("solid", fgColor=HDR_COLOR)
        c.alignment = Alignment(horizontal="center")
    for row_i, (_, row) in enumerate(sub.iterrows(), 2):
        ws2.cell(row=row_i, column=1, value=row['Period'])
        ws2.cell(row=row_i, column=2, value=row['CPI_Index'])
        ws2.cell(row=row_i, column=3, value=row['MoM_Pct'])
        ws2.cell(row=row_i, column=4, value=row['YoY_Pct'])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def get_signal(pct):
    if abs(pct) >= 4.0:
        return "🔴", "#FEE2E2", "Strong Review Recommended", "badge"
    elif abs(pct) >= 2.0:
        return "🟡", "#FEF9C3", "Selective Review Warranted", "badge-yellow"
    else:
        return "🟢", "#DCFCE7", "Stable — Continue Monitoring", "badge-green"

def render_signal_box(pct, sv, ev, start_period, end_period, city, latest_yoy, frequency="Monthly"):
    icon, bg, title, badge_cls = get_signal(pct)
    if abs(pct) >= 4.0:
        desc = f"CPI has risen <strong>{pct:+.2f}%</strong> ({start_period}→{end_period}), exceeding the 4% threshold. This supports a formal pricing adjustment proposal."
        badge_label = "≥ 4% Threshold"
    elif abs(pct) >= 2.0:
        desc = f"CPI has moved <strong>{pct:+.2f}%</strong> ({start_period}→{end_period}). Consider selective rate adjustments for CPI-sensitive categories."
        badge_label = "2–4% Range"
    else:
        desc = f"CPI has moved <strong>{pct:+.2f}%</strong> ({start_period}→{end_period}), within acceptable tolerance. No immediate pricing action required."
        badge_label = "< 2% Stable"

    freq_suffix = " (Quarterly Series)" if frequency == "Quarterly" else ""
    formula_note = f"Formula: (({ev:.2f} − {sv:.2f}) ÷ {sv:.2f}) × 100 = <strong>{pct:+.4f}%</strong> &nbsp;|&nbsp; ABS Official YoY{freq_suffix}: <strong>{latest_yoy}%</strong>"

    st.markdown(f"""
    <div class="signal-box" style="background:{bg};">
        <div class="signal-icon">{icon}</div>
        <div>
            <div class="signal-title">{title} <span class="badge {badge_cls}">{badge_label}</span></div>
            <div class="signal-desc">{desc}</div>
            <div class="signal-formula">{formula_note}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

def render_kpi_row(calc, city_df, city, start_period, end_period, frequency="Monthly"):
    pct  = calc['pct_change']
    mv   = calc['movement']
    sv   = calc['start_val']
    ev   = calc['end_val']
    latest_yoy = city_df.iloc[-1]['YoY_Pct']
    latest_mom = city_df.iloc[-1]['MoM_Pct']
    latest_period = city_df.iloc[-1]['Period']
    prev_period   = city_df.iloc[-2]['Period']
    n_periods = len(city_df[(city_df['Date'] >= city_df[city_df['Period'] == start_period]['Date'].values[0]) &
                             (city_df['Date'] <= city_df[city_df['Period'] == end_period]['Date'].values[0])])
    up = pct > 0
    pct_cls = "kpi-red" if up else "kpi-green"
    arrow   = "▲" if up else "▼"
    mom_label = "QoQ Change" if frequency == "Quarterly" else "Latest MoM Change"
    period_unit = "quarters" if frequency == "Quarterly" else "months"
    freq_pill = f'<span class="series-pill pill-quarterly">Quarterly</span>' if frequency == "Quarterly" \
                else f'<span class="series-pill pill-monthly">Monthly</span>'

    st.markdown(f"""
    <div class="kpi-wrap">
      <div class="kpi">
        <div class="kpi-label">Series {freq_pill}</div>
        <div class="kpi-val" style="font-size:1rem;color:#555;">{city}</div>
        <div class="kpi-sub">{start_period} → {end_period}</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">CPI Start Value</div>
        <div class="kpi-val">{sv:.2f}</div>
        <div class="kpi-sub">{start_period}</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">CPI End Value</div>
        <div class="kpi-val">{ev:.2f}</div>
        <div class="kpi-sub">{end_period}</div>
      </div>
      <div class="kpi {'kpi-red' if mv > 0 else 'kpi-green'}">
        <div class="kpi-label">Index Movement</div>
        <div class="kpi-val">{mv:+.2f}</div>
        <div class="kpi-sub">pts over {n_periods} {period_unit}</div>
      </div>
      <div class="kpi {pct_cls}">
        <div class="kpi-label">% Change</div>
        <div class="kpi-val">{arrow} {abs(pct):.2f}%</div>
        <div class="kpi-sub">{start_period} → {end_period}</div>
      </div>
      <div class="kpi {'kpi-red' if (latest_yoy or 0)>0 else 'kpi-green'}">
        <div class="kpi-label">Latest YoY (ABS Official)</div>
        <div class="kpi-val">{'▲' if (latest_yoy or 0)>0 else '▼'} {abs(latest_yoy or 0):.1f}%</div>
        <div class="kpi-sub">{latest_period} vs same period prior yr</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">{mom_label}</div>
        <div class="kpi-val">{'▲' if (latest_mom or 0)>0 else '▼'} {abs(latest_mom or 0):.1f}%</div>
        <div class="kpi-sub">{latest_period} vs {prev_period}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

def filter_city_df(city_df, start_period, end_period):
    start_date = city_df[city_df['Period'] == start_period]['Date'].values[0]
    end_date   = city_df[city_df['Period'] == end_period]['Date'].values[0]
    return city_df[(city_df['Date'] >= start_date) & (city_df['Date'] <= end_date)].copy()

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    df_monthly_bundled = load_bundled_data()
    latest_label = df_monthly_bundled['Period'].iloc[-1] if df_monthly_bundled is not None else "—"

    st.markdown(f"""
    <div class="ap-header">
        <div class="ap-logo">📮</div>
        <div>
            <h1>CPI Calculator <span style="font-weight:300;font-size:1.1rem;">| Australia Post</span></h1>
            <div class="sub">Data & Pricing Executive Tool &nbsp;·&nbsp; ABS 6401.0 Monthly + Quarterly CPI &nbsp;·&nbsp; Base: Sep-2025 = 100.00 &nbsp;·&nbsp; Latest Monthly: <strong>{latest_label}</strong></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── SIDEBAR ──────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown('<p class="sec-title">Series Mode</p>', unsafe_allow_html=True)
        series_mode = st.radio(
            "",
            ["📅 Monthly only", "📆 Quarterly only", "📊 Monthly + Quarterly (Dual)"],
            index=0,
            label_visibility="collapsed",
            help="Monthly = ABS 640101 | Quarterly = ABS 640106"
        )
        use_monthly   = "Monthly"   in series_mode
        use_quarterly = "Quarterly" in series_mode
        use_dual      = "Dual"      in series_mode

        st.markdown('<br><p class="sec-title">Data Sources</p>', unsafe_allow_html=True)

        df_monthly = df_quarterly = None

        # Monthly
        if use_monthly or use_dual:
            with st.expander("📅 Monthly Data (640101)", expanded=True):
                m_source = st.radio("", ["Bundled (Mar-2026)", "Upload new file"],
                                    key="m_src", label_visibility="collapsed")
                if m_source == "Upload new file":
                    m_up = st.file_uploader("Upload 640101.xlsx", type=["xlsx"], key="m_up")
                    if m_up:
                        _df_m_tmp = try_load_upload(m_up)
                        if _df_m_tmp is not None:
                            # Defensive: Frequency column may be missing if old parser is deployed
                            if 'Frequency' not in _df_m_tmp.columns:
                                if len(_df_m_tmp) >= 2:
                                    delta = (_df_m_tmp['Date'].iloc[1] - _df_m_tmp['Date'].iloc[0]).days
                                    _df_m_tmp['Frequency'] = 'Quarterly' if delta > 60 else 'Monthly'
                                else:
                                    _df_m_tmp['Frequency'] = 'Monthly'
                                if 'Period' not in _df_m_tmp.columns:
                                    _df_m_tmp['Period'] = _df_m_tmp['Date'].dt.strftime('%b-%Y')
                            freq_detected_m = _df_m_tmp['Frequency'].iloc[0]
                            if freq_detected_m != "Monthly":
                                st.error(f"⚠️ Wrong file — detected **{freq_detected_m}** data. Please upload **640101.xlsx** (Table 1 — Monthly) from ABS.")
                                df_monthly = df_monthly_bundled
                            else:
                                df_monthly = _df_m_tmp
                                st.success(f"✓ {len(df_monthly)} months ({df_monthly['Period'].iloc[0]} → {df_monthly['Period'].iloc[-1]}) · Monthly")
                else:
                    df_monthly = df_monthly_bundled

        # Quarterly
        if use_quarterly or use_dual:
            with st.expander("📆 Quarterly Data (Tables 17/18)", expanded=True):
                q_up = st.file_uploader("Upload quarterly CPI file (Tables 17/18)", type=["xlsx"], key="q_up",
                                        help="Download from ABS → CPI Australia → Table 6")
                if q_up:
                    _df_q_tmp = try_load_upload(q_up)
                    if _df_q_tmp is not None:
                        # Defensive: Frequency column may be missing if old parser is deployed
                        if 'Frequency' not in _df_q_tmp.columns:
                            # Infer from date spacing
                            if len(_df_q_tmp) >= 2:
                                delta = (_df_q_tmp['Date'].iloc[1] - _df_q_tmp['Date'].iloc[0]).days
                                _df_q_tmp['Frequency'] = 'Quarterly' if delta > 60 else 'Monthly'
                            else:
                                _df_q_tmp['Frequency'] = 'Monthly'
                            # Also add Period column if missing
                            if 'Period' not in _df_q_tmp.columns:
                                _df_q_tmp['Period'] = _df_q_tmp['Date'].dt.strftime('%b-%Y')
                        freq_detected = _df_q_tmp['Frequency'].iloc[0]
                        if freq_detected != "Quarterly":
                            st.error(
                                f"⚠️ Wrong file — detected **{freq_detected}** series. "
                                f"Please upload the quarterly CPI file (Tables 17 or 18) from the ABS data downloads page."
                            )
                            df_quarterly = None
                        else:
                            df_quarterly = _df_q_tmp
                            st.success(f"✓ {len(df_quarterly)} quarters ({df_quarterly['Period'].iloc[0]} → {df_quarterly['Period'].iloc[-1]}) · Quarterly")
                else:
                    st.markdown("""<div class="abs-note">
                    Download the <strong>quarterly CPI file (Tables 17 &amp; 18)</strong> from the
                    <a href="https://www.abs.gov.au/statistics/economy/price-indexes-and-inflation/consumer-price-index-australia/latest-release#data-downloads" target="_blank">ABS data downloads</a>
                    section — look for <strong>&ldquo;Table 17&rdquo;</strong> or <strong>&ldquo;Table 18&rdquo;</strong> in the download list.
                    <em>Note: 640106.xlsx is an expenditure group table — it will not work here.</em>
                    </div>""", unsafe_allow_html=True)

        # Validate we have what we need
        if (use_monthly or use_dual) and df_monthly is None:
            df_monthly = df_monthly_bundled
        if (use_quarterly or use_dual) and df_quarterly is None:
            if use_quarterly and not use_dual:
                st.warning("⚠️ Upload the correct quarterly ABS file (640106.xlsx — Table 6) to use this mode.")
                st.stop()

        st.markdown('<br><p class="sec-title">Filters</p>', unsafe_allow_html=True)
        city = st.selectbox("📍 Region", CITIES, index=0)

        # Period selectors depend on available data
        if use_monthly and not use_dual:
            active_df = df_monthly
        elif use_quarterly and not use_dual:
            if df_quarterly is None:
                st.stop()
            active_df = df_quarterly
        else:
            active_df = df_monthly  # dual uses monthly periods for the selector

        periods = active_df["Period"].tolist() if active_df is not None else []

        if not periods:
            st.error("No data available. Check your file uploads.")
            st.stop()

        is_quarterly_mode = (use_quarterly and not use_dual) and (active_df is not None and active_df["Frequency"].iloc[0] == "Quarterly")

        if is_quarterly_mode:
            # Build structured year → quarter picker
            import re
            q_df = active_df.copy()
            q_df["Year"]    = q_df["Period"].str.extract(r"Q\d-(\d{4})").astype(int)
            q_df["Quarter"] = q_df["Period"].str.extract(r"(Q\d)-")
            years_avail = sorted(q_df["Year"].unique().tolist())

            st.markdown("**Start Period**")
            sc1, sc2 = st.columns(2)
            with sc1:
                s_year = st.selectbox("Year", years_avail, index=0, key="s_year")
            with sc2:
                s_qtrs = q_df[q_df["Year"] == s_year]["Quarter"].tolist()
                s_qtr  = st.selectbox("Quarter", s_qtrs, index=0, key="s_qtr")
            start_period = f"{s_qtr}-{s_year}"

            st.markdown("**End Period**")
            ec1, ec2 = st.columns(2)
            with ec1:
                e_year = st.selectbox("Year", years_avail, index=len(years_avail)-1, key="e_year")
            with ec2:
                e_qtrs = q_df[q_df["Year"] == e_year]["Quarter"].tolist()
                e_qtr  = st.selectbox("Quarter", e_qtrs, index=len(e_qtrs)-1, key="e_qtr")
            end_period = f"{e_qtr}-{e_year}"

        else:
            col_s, col_e = st.columns(2)
            with col_s:
                start_period = st.selectbox("Start", periods,
                                            index=max(0, len(periods)-13),
                                            help="Base period for comparison")
            with col_e:
                end_period = st.selectbox("End", periods,
                                          index=len(periods)-1,
                                          help="Comparison end period")

        st.markdown('<br><p class="sec-title">Display Options</p>', unsafe_allow_html=True)
        show_all_regions = st.checkbox("Overlay all regions on index chart", value=False)
        show_yoy_chart   = st.checkbox("Show YoY % chart", value=True)
        show_pricing     = st.checkbox("Pricing guidance panel", value=True)

    # ── VALIDATION ───────────────────────────────────────────────────────────
    if start_period not in periods or end_period not in periods:
        st.error("⚠️ Selected period not found in data. Check your file.")
        st.stop()
    s_idx = periods.index(start_period)
    e_idx = periods.index(end_period)
    if s_idx >= e_idx:
        st.error("⚠️ Start period must be before End period.")
        st.stop()

    # ═══════════════════════════════════════════════════════════════════════
    # DUAL MODE
    # ═══════════════════════════════════════════════════════════════════════
    if use_dual:
        if df_quarterly is None:
            st.info("📆 Upload the quarterly file (640106.xlsx) in the sidebar to activate dual-series view.")
            # Fall through to monthly-only display
            use_dual = False
            use_monthly = True

    if use_dual and df_quarterly is not None:
        _render_dual(df_monthly, df_quarterly, city, start_period, end_period,
                     show_all_regions, show_yoy_chart, show_pricing)
        return

    # ═══════════════════════════════════════════════════════════════════════
    # SINGLE SERIES MODE (Monthly or Quarterly)
    # ═══════════════════════════════════════════════════════════════════════
    df = df_monthly if use_monthly else df_quarterly
    frequency = df['Frequency'].iloc[0] if df is not None else "Monthly"
    _render_single(df, city, start_period, end_period,
                   show_all_regions, show_yoy_chart, show_pricing, frequency)


# ─── SINGLE SERIES RENDERER ──────────────────────────────────────────────────

def _render_single(df, city, start_period, end_period,
                   show_all_regions, show_yoy_chart, show_pricing, frequency="Monthly"):
    calc     = calc_custom_change(df, city, start_period, end_period)
    city_df  = get_city_df(df, city)
    pct, sv, ev = calc['pct_change'], calc['start_val'], calc['end_val']
    latest_yoy = city_df.iloc[-1]['YoY_Pct']
    pop_label  = "QoQ" if frequency == "Quarterly" else "MoM"
    series_color = QUARTERLY_COLOR if frequency == "Quarterly" else MONTHLY_COLOR

    render_kpi_row(calc, city_df, city, start_period, end_period, frequency)

    if show_pricing:
        st.markdown('<p class="sec-title">Pricing Signal</p>', unsafe_allow_html=True)
        render_signal_box(pct, sv, ev, start_period, end_period, city, latest_yoy, frequency)

    sub = filter_city_df(city_df, start_period, end_period)

    chart_col1, chart_col2 = (st.columns(2) if show_yoy_chart else (st.container(), None))

    with (chart_col1 if show_yoy_chart else st):
        st.markdown('<p class="sec-title">CPI Index Trend</p>', unsafe_allow_html=True)
        fig = go.Figure()
        if show_all_regions:
            colors_map = {
                'Australia': MONTHLY_COLOR, 'Sydney': '#1E40AF', 'Melbourne': '#15803D',
                'Brisbane': '#92400E', 'Adelaide': '#7C3AED', 'Perth': '#0891B2',
                'Hobart': '#B45309', 'Darwin': '#BE185D', 'Canberra': '#374151'
            }
            for c in CITIES:
                c_sub = filter_city_df(get_city_df(df, c), start_period, end_period)
                fig.add_trace(go.Scatter(
                    x=c_sub['Period'], y=c_sub['CPI_Index'], name=c,
                    line=dict(width=2.5 if c == city else 1.5, color=colors_map[c],
                              dash='solid' if c == city else 'dot'),
                    mode="lines",
                ))
            chart_title = f"All Regions — CPI Index ({frequency})"
        else:
            fig.add_trace(go.Scatter(
                x=sub['Period'], y=sub['CPI_Index'], name=city,
                line=dict(color=series_color, width=2.5),
                mode="lines+markers",
                marker=dict(size=7, color=series_color),
                fill="tozeroy", fillcolor=f"rgba({int(series_color[1:3],16)},{int(series_color[3:5],16)},{int(series_color[5:7],16)},0.06)",
            ))
            fig.add_annotation(x=start_period, y=sv,
                text=f"<b>{sv:.2f}</b><br>{start_period}",
                showarrow=True, arrowhead=2, ax=-40, ay=-40,
                font=dict(size=10, color="#555"), bgcolor="white", bordercolor="#ddd", borderwidth=1)
            fig.add_annotation(x=end_period, y=ev,
                text=f"<b>{ev:.2f}</b><br>{end_period}",
                showarrow=True, arrowhead=2, ax=40, ay=-40,
                font=dict(size=10, color=series_color), bgcolor="white", bordercolor=series_color, borderwidth=1)
            chart_title = f"{city} — CPI Index ({frequency}, Base: Sep-2025 = 100)"

        fig.add_hline(y=100, line_dash="dash", line_color="#ccc", line_width=1,
                      annotation_text="Base: Sep-2025 = 100", annotation_position="right")
        fig.update_layout(
            title=dict(text=chart_title, font=dict(size=13, color="#111"), x=0),
            xaxis=dict(tickangle=-45, gridcolor="#f5f5f5"),
            yaxis=dict(title="Index", gridcolor="#f5f5f5"),
            plot_bgcolor="white", paper_bgcolor="white",
            hovermode="x unified", height=340,
            legend=dict(orientation="h", y=1.08, x=0),
            margin=dict(l=40, r=10, t=50, b=60),
            font=dict(family="IBM Plex Sans"),
        )
        st.plotly_chart(fig, use_container_width=True)

    if show_yoy_chart and chart_col2:
        with chart_col2:
            st.markdown(f'<p class="sec-title">YoY % Change (ABS Official)</p>', unsafe_allow_html=True)
            yoy_sub = sub.dropna(subset=['YoY_Pct'])
            if yoy_sub.empty:
                st.info("YoY data not available for this period range.")
            else:
                fig2 = go.Figure()
                bar_colors = [MONTHLY_COLOR if v >= 4 else ("#F59E0B" if v >= 2 else "#16A34A")
                              for v in yoy_sub['YoY_Pct']]
                fig2.add_trace(go.Bar(
                    x=yoy_sub['Period'], y=yoy_sub['YoY_Pct'],
                    marker_color=bar_colors, name="YoY %",
                    text=yoy_sub['YoY_Pct'].apply(lambda x: f"{x:.1f}%"),
                    textposition="outside", textfont=dict(size=10),
                ))
                fig2.add_hline(y=4.0, line_dash="dash", line_color=MONTHLY_COLOR, line_width=1.5,
                               annotation_text="4% review threshold", annotation_position="right",
                               annotation_font=dict(color=MONTHLY_COLOR, size=10))
                fig2.add_hline(y=2.0, line_dash="dot", line_color="#F59E0B", line_width=1,
                               annotation_text="2% monitor level", annotation_position="right",
                               annotation_font=dict(color="#F59E0B", size=9))
                fig2.update_layout(
                    title=dict(text=f"{city} — Annual CPI Change % ({frequency})", font=dict(size=13, color="#111"), x=0),
                    xaxis=dict(tickangle=-45, gridcolor="#f5f5f5"),
                    yaxis=dict(title="%", gridcolor="#f5f5f5"),
                    plot_bgcolor="white", paper_bgcolor="white",
                    hovermode="x unified", height=340,
                    margin=dict(l=40, r=10, t=50, b=60),
                    font=dict(family="IBM Plex Sans"),
                    showlegend=False,
                )
                st.plotly_chart(fig2, use_container_width=True)

    _render_table(sub, city_df, start_period, end_period, sv, frequency)
    _render_snapshot(df, city)
    _render_exports(city_df, city, start_period, end_period, calc, frequency)
    _render_footer(df, frequency)


# ─── DUAL SERIES RENDERER ────────────────────────────────────────────────────

def _render_dual(df_m, df_q, city, start_period, end_period,
                 show_all_regions, show_yoy_chart, show_pricing):
    """Dual monthly + quarterly comparison view."""
    city_m = get_city_df(df_m, city)
    city_q = get_city_df(df_q, city)

    calc_m = calc_custom_change(df_m, city, start_period, end_period)

    # For quarterly, find closest matching period range
    q_periods = df_q['Period'].tolist()
    q_start   = q_periods[0]
    q_end     = q_periods[-1]
    calc_q    = calc_custom_change(df_q, city, q_start, q_end)

    sub_m = filter_city_df(city_m, start_period, end_period)
    sub_q = city_q.copy()  # show all available quarterly data

    pct_m = calc_m['pct_change']
    pct_q = calc_q['pct_change']
    icon_m, _, _, _ = get_signal(pct_m)
    icon_q, _, _, _ = get_signal(pct_q)

    # ── DUAL HEADER ──
    st.markdown(f"""
    <div class="dual-header">
        <div style="font-size:1.4rem;">📊</div>
        <div class="dual-header-item">
            <div class="dual-header-label"><span class="series-pill pill-monthly">Monthly</span> {start_period} → {end_period}</div>
            <div class="dual-header-val dual-monthly">{icon_m} {pct_m:+.2f}%</div>
        </div>
        <div style="color:#ccc;font-size:1.5rem;">|</div>
        <div class="dual-header-item">
            <div class="dual-header-label"><span class="series-pill pill-quarterly">Quarterly</span> {q_start} → {q_end}</div>
            <div class="dual-header-val dual-quarterly">{icon_q} {pct_q:+.2f}%</div>
        </div>
        <div style="color:#ccc;font-size:1.5rem;">|</div>
        <div class="dual-header-item">
            <div class="dual-header-label">Region</div>
            <div class="dual-header-val" style="color:#374151;">{city}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── KPI ROWS ──
    col_m, col_q = st.columns(2)
    with col_m:
        st.markdown('<p class="sec-title">Monthly Series</p>', unsafe_allow_html=True)
        render_kpi_row(calc_m, city_m, city, start_period, end_period, "Monthly")
    with col_q:
        st.markdown('<p class="sec-title">Quarterly Series</p>', unsafe_allow_html=True)
        render_kpi_row(calc_q, city_q, city, q_start, q_end, "Quarterly")

    # ── PRICING SIGNALS ──
    if show_pricing:
        st.markdown('<p class="sec-title">Pricing Signals — Comparison</p>', unsafe_allow_html=True)
        sc1, sc2 = st.columns(2)
        with sc1:
            render_signal_box(pct_m, calc_m['start_val'], calc_m['end_val'],
                              start_period, end_period, city,
                              city_m.iloc[-1]['YoY_Pct'], "Monthly")
        with sc2:
            render_signal_box(pct_q, calc_q['start_val'], calc_q['end_val'],
                              q_start, q_end, city,
                              city_q.iloc[-1]['YoY_Pct'], "Quarterly")

    # ── COMBINED INDEX TREND CHART ──
    st.markdown('<p class="sec-title">CPI Index — Monthly vs Quarterly Overlay</p>', unsafe_allow_html=True)
    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=sub_m['Period'], y=sub_m['CPI_Index'],
        name="Monthly", line=dict(color=MONTHLY_COLOR, width=2.5),
        mode="lines+markers", marker=dict(size=5, color=MONTHLY_COLOR),
    ))
    fig.add_trace(go.Scatter(
        x=sub_q['Period'], y=sub_q['CPI_Index'],
        name="Quarterly", line=dict(color=QUARTERLY_COLOR, width=2.5, dash="dash"),
        mode="lines+markers", marker=dict(size=8, color=QUARTERLY_COLOR, symbol="diamond"),
    ))
    fig.add_hline(y=100, line_dash="dash", line_color="#ccc", line_width=1,
                  annotation_text="Base: Sep-2025 = 100", annotation_position="right")
    fig.update_layout(
        title=dict(text=f"{city} — CPI Index: Monthly (red) vs Quarterly (blue)", font=dict(size=13, color="#111"), x=0),
        xaxis=dict(tickangle=-45, gridcolor="#f5f5f5"),
        yaxis=dict(title="CPI Index", gridcolor="#f5f5f5"),
        plot_bgcolor="white", paper_bgcolor="white",
        hovermode="x unified", height=360,
        legend=dict(orientation="h", y=1.08, x=0),
        margin=dict(l=40, r=10, t=50, b=60),
        font=dict(family="IBM Plex Sans"),
    )
    st.plotly_chart(fig, use_container_width=True)

    # ── YoY COMPARISON CHART ──
    if show_yoy_chart:
        st.markdown('<p class="sec-title">YoY % Change — Monthly vs Quarterly</p>', unsafe_allow_html=True)
        yoy_m = sub_m.dropna(subset=['YoY_Pct'])
        yoy_q = sub_q.dropna(subset=['YoY_Pct'])

        fig3 = go.Figure()
        if not yoy_m.empty:
            fig3.add_trace(go.Bar(
                x=yoy_m['Period'], y=yoy_m['YoY_Pct'],
                name="Monthly YoY", marker_color=MONTHLY_COLOR, opacity=0.8,
                text=yoy_m['YoY_Pct'].apply(lambda x: f"{x:.1f}%"),
                textposition="outside", textfont=dict(size=9),
            ))
        if not yoy_q.empty:
            fig3.add_trace(go.Bar(
                x=yoy_q['Period'], y=yoy_q['YoY_Pct'],
                name="Quarterly YoY", marker_color=QUARTERLY_COLOR, opacity=0.8,
                text=yoy_q['YoY_Pct'].apply(lambda x: f"{x:.1f}%"),
                textposition="outside", textfont=dict(size=9),
            ))
        fig3.add_hline(y=4.0, line_dash="dash", line_color=MONTHLY_COLOR, line_width=1.5,
                       annotation_text="4% threshold", annotation_position="right",
                       annotation_font=dict(color=MONTHLY_COLOR, size=10))
        fig3.add_hline(y=2.0, line_dash="dot", line_color="#F59E0B", line_width=1,
                       annotation_text="2% monitor", annotation_position="right",
                       annotation_font=dict(color="#F59E0B", size=9))
        fig3.update_layout(
            title=dict(text=f"{city} — Annual CPI Change: Monthly vs Quarterly", font=dict(size=13, color="#111"), x=0),
            xaxis=dict(tickangle=-45, gridcolor="#f5f5f5"),
            yaxis=dict(title="%", gridcolor="#f5f5f5"),
            plot_bgcolor="white", paper_bgcolor="white",
            hovermode="x unified", height=360, barmode="group",
            legend=dict(orientation="h", y=1.08, x=0),
            margin=dict(l=40, r=10, t=50, b=60),
            font=dict(family="IBM Plex Sans"),
        )
        st.plotly_chart(fig3, use_container_width=True)

    # ── SIDE-BY-SIDE TABLES ──
    st.markdown('<p class="sec-title">Detailed Results</p>', unsafe_allow_html=True)
    tc1, tc2 = st.columns(2)
    with tc1:
        st.caption("📅 Monthly")
        _render_table(sub_m, city_m, start_period, end_period, calc_m['start_val'], "Monthly", compact=True)
    with tc2:
        st.caption("📆 Quarterly")
        _render_table(sub_q, city_q, q_start, q_end, calc_q['start_val'], "Quarterly", compact=True)

    # ── SNAPSHOT ──
    _render_snapshot(df_m, city)

    # ── EXPORTS ──
    st.markdown('<p class="sec-title">Export</p>', unsafe_allow_html=True)
    ec1, ec2, ec3, ec4 = st.columns([1, 1, 1, 2])
    with ec1:
        csv_m = sub_m[['Period','CPI_Index','MoM_Pct','YoY_Pct']].to_csv(index=False).encode("utf-8")
        st.download_button("⬇ Monthly CSV", data=csv_m,
                           file_name=f"CPI_Monthly_{city}_{start_period}_{end_period}.csv",
                           mime="text/csv")
    with ec2:
        xlsx_m = build_excel_report(city_m, city, start_period, end_period, calc_m, "Monthly")
        st.download_button("⬇ Monthly Excel", data=xlsx_m,
                           file_name=f"CPI_Monthly_{city}_{start_period}_{end_period}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with ec3:
        csv_q = sub_q[['Period','CPI_Index','MoM_Pct','YoY_Pct']].to_csv(index=False).encode("utf-8")
        st.download_button("⬇ Quarterly CSV", data=csv_q,
                           file_name=f"CPI_Quarterly_{city}_{q_start}_{q_end}.csv",
                           mime="text/csv")

    _render_footer(df_m, "Monthly + Quarterly")


# ─── SHARED RENDER HELPERS ───────────────────────────────────────────────────

def _render_table(sub, city_df, start_period, end_period, sv, frequency, compact=False):
    pop_label = "QoQ (%)" if frequency == "Quarterly" else "MoM (%)"
    sub_display = sub.copy()
    sub_display['vs_Start_%'] = ((sub_display['CPI_Index'] - sv) / sv * 100).round(2)
    display_df = sub_display[['Period', 'CPI_Index', 'MoM_Pct', 'YoY_Pct', 'vs_Start_%']].copy()
    display_df.columns = ['Period', 'CPI Index', pop_label, 'YoY (%)', f'vs Start (%)']

    def color_pct(val):
        if pd.isna(val) or not isinstance(val, (int, float)): return ''
        if val > 0: return 'color: #C8102E; font-weight: 600'
        if val < 0: return 'color: #15803d; font-weight: 600'
        return ''

    styled = (display_df.style
        .format({'CPI Index': '{:.2f}',
                 pop_label: lambda x: f'{x:+.1f}%' if pd.notna(x) else '—',
                 'YoY (%)': lambda x: f'{x:+.1f}%' if pd.notna(x) else '—',
                 'vs Start (%)': '{:+.2f}%'})
        .map(color_pct, subset=[pop_label, 'YoY (%)', 'vs Start (%)'])
    )
    if not compact:
        st.markdown('<p class="sec-title">Detailed Results Table</p>', unsafe_allow_html=True)
    st.dataframe(styled, use_container_width=True, hide_index=True)


def _render_snapshot(df, city):
    with st.expander("📊 All Regions Snapshot — Latest Month"):
        latest_row = df.iloc[-1]
        latest_period = latest_row['Period']
        snap_data = []
        for c in CITIES:
            idx_val = latest_row[f'{c}_Index']
            yoy_val = latest_row[f'{c}_YoY']
            mom_val = latest_row[f'{c}_MoM']
            signal = "🔴" if (yoy_val or 0) >= 4 else ("🟡" if (yoy_val or 0) >= 2 else "🟢")
            snap_data.append({
                'Region': c,
                'CPI Index': f"{idx_val:.2f}" if idx_val else '—',
                'YoY % (Official)': f"{yoy_val:+.1f}%" if yoy_val else '—',
                'MoM %': f"{mom_val:+.1f}%" if mom_val else '—',
                'Signal': signal,
            })
        st.caption(f"Reference period: **{latest_period}** — Source: ABS 6401.0")
        st.dataframe(pd.DataFrame(snap_data), use_container_width=True, hide_index=True)


def _render_exports(city_df, city, start_period, end_period, calc, frequency):
    st.markdown('<p class="sec-title">Export</p>', unsafe_allow_html=True)
    sub = filter_city_df(city_df, start_period, end_period)
    pop_label = "QoQ (%)" if frequency == "Quarterly" else "MoM (%)"
    sub_display = sub[['Period', 'CPI_Index', 'MoM_Pct', 'YoY_Pct']].copy()
    sub_display.columns = ['Period', 'CPI Index', pop_label, 'YoY (%)']
    ec1, ec2, ec3 = st.columns([1, 1, 3])
    with ec1:
        csv_out = sub_display.to_csv(index=False).encode("utf-8")
        st.download_button("⬇ Download CSV", data=csv_out,
                           file_name=f"CPI_{city}_{start_period}_{end_period}.csv",
                           mime="text/csv")
    with ec2:
        xlsx_out = build_excel_report(city_df, city, start_period, end_period, calc, frequency)
        st.download_button("⬇ Download Excel Report", data=xlsx_out,
                           file_name=f"CPI_Report_{city}_{start_period}_{end_period}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _render_footer(df, frequency_label):
    st.markdown("---")
    latest_period = df['Period'].iloc[-1]
    st.markdown(f"""
    <div style="font-size:0.75rem;color:#bbb;text-align:center;padding:6px 0;">
        Australia Post · CPI Calculator v3.0 · Data & Pricing Executive Tool ·
        Source: <a href="https://www.abs.gov.au/statistics/economy/price-indexes-and-inflation/consumer-price-index-australia/latest-release" target="_blank" style="color:#C8102E;">ABS 6401.0 ({frequency_label}, {latest_period})</a> ·
        Base Sep-2025 = 100
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()

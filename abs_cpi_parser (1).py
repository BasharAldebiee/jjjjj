"""
ABS 6401.0 CPI Data Parser
Handles Monthly (Table 1, 640101.xlsx) and Quarterly (Tables 17/18) workbooks.

Monthly file: 640101.xlsx — Data1 sheet, cols 1-9 index, 10-18 YoY, 19-27 MoM
Quarterly file: From ABS downloads section "Tables 17 and 18" — same Data1 layout
                but Frequency row says "Quarter" and dates are quarterly.

Auto-detects frequency from:
  1. Frequency metadata row (row 5): "Month" → Monthly, "Quarter" → Quarterly
  2. Fallback: date gap between first two rows (>60 days → Quarterly)
"""

import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

CITIES = ['Australia', 'Sydney', 'Melbourne', 'Brisbane', 'Adelaide',
          'Perth', 'Hobart', 'Darwin', 'Canberra']

CITY_COL_MAP = {
    'index': list(range(1, 10)),
    'yoy':   list(range(10, 19)),
    'mom':   list(range(19, 28)),
}


def _detect_frequency_from_rows(rows) -> str:
    """Detect Monthly vs Quarterly. Checks metadata row 5 first, then date gap."""
    # Row 5 (index 4) = Frequency metadata
    if len(rows) > 4:
        freq_row = rows[4]
        for val in freq_row[1:]:
            if isinstance(val, str):
                v = val.strip().lower()
                if 'quarter' in v:
                    return 'Quarterly'
                if 'month' in v:
                    return 'Monthly'

    # Fallback: check date gap between first two data rows (row index 10+)
    data_rows = [r for r in rows[10:] if r[0] is not None]
    if len(data_rows) >= 2:
        try:
            d1 = pd.Timestamp(data_rows[0][0])
            d2 = pd.Timestamp(data_rows[1][0])
            delta = (d2 - d1).days
            return 'Quarterly' if delta > 60 else 'Monthly'
        except Exception:
            pass

    return 'Monthly'


def _find_data_sheet(wb) -> str:
    """Find the data sheet — ABS uses 'Data1' for most tables."""
    preferred = ['Data1', 'Data 1', 'Table 17', 'Table 18', 'Sheet1']
    for name in preferred:
        if name in wb.sheetnames:
            return name
    # Fall back to first non-index sheet
    for name in wb.sheetnames:
        if name.lower() not in ('index', 'contents', 'notes', 'glossary'):
            return name
    return wb.sheetnames[0]


def load_abs_file(filepath_or_bytes) -> pd.DataFrame:
    """
    Parse any ABS 6401.0 CPI Time Series Excel file.

    Works with:
      - 640101.xlsx  (Monthly, Table 1)
      - Tables 17/18 quarterly download from ABS data downloads section

    Returns DataFrame with columns:
      Date, Period, Frequency, {City}_Index, {City}_YoY, {City}_MoM
    """
    if isinstance(filepath_or_bytes, bytes):
        filepath_or_bytes = BytesIO(filepath_or_bytes)

    wb = load_workbook(filepath_or_bytes, read_only=True, data_only=True)
    sheet_name = _find_data_sheet(wb)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    frequency = _detect_frequency_from_rows(rows)

    records = []
    for r in rows[10:]:
        dt = r[0]
        if dt is None:
            continue
        try:
            ts = pd.Timestamp(dt)
        except Exception:
            continue
        record = {'Date': ts, 'Frequency': frequency}
        for i, city in enumerate(CITIES):
            idx_i = CITY_COL_MAP['index'][i]
            yoy_i = CITY_COL_MAP['yoy'][i]
            mom_i = CITY_COL_MAP['mom'][i]
            record[f'{city}_Index'] = _safe_float(r[idx_i] if len(r) > idx_i else None)
            record[f'{city}_YoY']   = _safe_float(r[yoy_i] if len(r) > yoy_i else None)
            record[f'{city}_MoM']   = _safe_float(r[mom_i] if len(r) > mom_i else None)
        records.append(record)

    if not records:
        raise ValueError(f"No data rows found in sheet '{sheet_name}'. Check the file format.")

    df = pd.DataFrame(records).sort_values('Date').reset_index(drop=True)

    if frequency == 'Quarterly':
        df['Period'] = df['Date'].apply(_quarter_label)
    else:
        df['Period'] = df['Date'].dt.strftime('%b-%Y')

    return df


def _quarter_label(dt: pd.Timestamp) -> str:
    q = (dt.month - 1) // 3 + 1
    return f"Q{q}-{dt.year}"


def _safe_float(val):
    try:
        return float(val) if val is not None else None
    except (TypeError, ValueError):
        return None


def get_city_df(df: pd.DataFrame, city: str) -> pd.DataFrame:
    cols = ['Date', 'Period', 'Frequency', f'{city}_Index', f'{city}_YoY', f'{city}_MoM']
    out = df[cols].copy()
    out.columns = ['Date', 'Period', 'Frequency', 'CPI_Index', 'YoY_Pct', 'MoM_Pct']
    return out


def calc_custom_change(df: pd.DataFrame, city: str, start_period: str, end_period: str) -> dict:
    city_df = get_city_df(df, city)
    s = city_df[city_df['Period'] == start_period]['CPI_Index'].values
    e = city_df[city_df['Period'] == end_period]['CPI_Index'].values
    if len(s) == 0 or len(e) == 0:
        return {}
    start_val, end_val = float(s[0]), float(e[0])
    pct = ((end_val - start_val) / start_val) * 100
    return {
        'start_period': start_period,
        'end_period': end_period,
        'start_val': start_val,
        'end_val': end_val,
        'movement': end_val - start_val,
        'pct_change': pct,
    }
